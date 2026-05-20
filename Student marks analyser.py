import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from sklearn.linear_model import LinearRegression
from openpyxl import Workbook

# Load dataset
students = []

n = int(input("How many students? "))

for i in range(n):
    print(f"\nEnter details for student {i+1}")

    name = input("Name: ")
    math = int(input("Math marks: "))
    science = int(input("Science marks: "))
    english = int(input("English marks: "))
    attendance = int(input("Attendance: "))
    study_hours = int(input("Study hours: "))

    students.append({
        "Name": name,
        "Math": math,
        "Science": science,
        "English": english,
        "Attendance": attendance,
        "StudyHours": study_hours
    })

# Create DataFrame
df = pd.DataFrame(students)

# Calculate average
df["Average"] = (df["Math"] + df["Science"] + df["English"]) / 3

print(df)

df.to_excel("report.xlsx", index=False)

print("Excel report created")
def grade(avg):
    if avg >= 90:
        return "A(Pass)"
    elif avg >= 75:
        return "B(Pass)"
    elif avg >= 50:
        return "C(Pass)"
    else:
        return "Fail"

df["Grade"] = df["Average"].apply(grade)
print("Student Data:\n")


# Top performer
df["Rank"] = df["Average"].rank(ascending=False)
topper = df.loc[df["Average"].idxmax()]
print("\nTop Performer:")
print(topper["Name"])
print(df)
#Subject Toppers
print("Math Topper:")
print(df.loc[df["Math"].idxmax()]["Name"])
print("English Topper:")
print(df.loc[df["English"].idxmax()]["Name"])
print("Science Topper:")
print(df.loc[df["Science"].idxmax()]["Name"])
# Weak Students
weak_students = df[df["Average"] < 50]
print(weak_students)
# Average marks chart
plt.figure(figsize=(8,5))
sns.barplot(x="Name", y="Average", data=df)

plt.title("Average Marks of Students")
plt.xlabel("Students")
plt.ylabel("Average Marks")

plt.savefig("charts/average_marks.png")
plt.show()

# Study hours vs Average
plt.figure(figsize=(8,5))
sns.scatterplot(x="StudyHours", y="Average", data=df)

plt.title("Study Hours vs Average Marks")

plt.savefig("charts/study_vs_marks.png")
plt.show()
#grade distribution
grade_counts = df["Grade"].value_counts()

plt.pie(grade_counts, labels=grade_counts.index, autopct='%1.1f%%')
plt.title("Grade Distribution")

plt.show()
 #Subject performance
sns.heatmap(df[["Math","Science","English"]], annot=True)

plt.title("Subject Heatmap")
plt.show()
#Export Reports
from openpyxl import Workbook

wb = Workbook()
ws = wb.active

ws["A1"] = "Name"
ws["B1"] = "Marks"
df.to_excel("report.xlsx", index=False)
#Attendance Analysis
sns.scatterplot(x="Attendance", y="Average", data=df)

